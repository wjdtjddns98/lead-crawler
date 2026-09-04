"""기존 임포트 행 업종 백필 ② — 원본 엑셀 '구분' 컬럼 재사용 (2026-08-19 PO 지시).

임포트 시드(importer.py)는 dedup 목적이라 원본 엑셀의 '구분'(업종) 컬럼을 버렸다.
원본(바탕화면 DB/DB1/DB2)을 다시 읽어 canonical_key 로 미승격·미분류 import 행과
조인(실측 77%)하고, 구분 **표기값 단위로 1회** 택소노미 라벨을 정해 일괄 적용한다 —
행당 홈페이지 LLM(backfill_import_industry.py) 대비 호출을 표기 종수로 줄인다.

표기값 → 라벨 결정(우선순위):
  1. 수동 별칭(투자사 계열·연기금 등 자명한 값 + 비업종 표기의 명시적 제외)
  2. 이미 택소노미 라벨과 동일한 표기(그대로)
  3. LLM 값 단위 분류 — nps-map-industries(storage/nps.map_industry_codes) 선례와 동일:
     classifier.classify(업종표기 + 대표 회사명 샘플, 홈페이지 없음). abstain=스킵.
거래소 보드명·투자 라운드 등 비업종 표기는 1·3에서 걸러져 미분류로 남는다(→ 행당
스크립트/승격 분류가 후속). 멱등: 라벨이 채워진 행은 다음 실행 대상에서 빠진다.

사용:  python scripts/backfill_import_gubun.py [--limit-values 0] [--no-llm]

ponytail: 1회성 백필 — 매핑을 테이블로 저장하지 않는다(중단 시 미처리 표기만 재과금,
표기당 5원). 상시화되면 ksic_industry_map 류 캐시로 승격.
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path

from sqlalchemy import text

from leadcrawler.config import Settings
from leadcrawler.cost_ledger import CostLedger
from leadcrawler.dedup import canonical_key
from leadcrawler.enrich.industry_classify import build_classifier
from leadcrawler.sources.taxonomy import INDUSTRY_TAXONOMY
from leadcrawler.storage.db import get_sessionmaker

DIRS = [Path(r"C:\Users\WSCOPY\Desktop") / d for d in ("DB", "DB1", "DB2")]
# importer._HEADER_ALIASES 와 동일 + 이번 목적의 '구분'/'업종'(임포터는 버리는 컬럼).
_ALIASES = {"국가": "country", "업체명": "name", "회사명": "name", "사이트": "domain",
            "홈페이지": "domain", "site": "domain", "도메인": "domain",
            "구분": "gubun", "업종": "gubun"}

# 수동 별칭 — 자명한 값만(추측성 매핑 금지, 애매하면 LLM/제외). None=비업종 표기(제외).
_HAND: dict[str, str | None] = {
    "자산운용사": "증권·자산운용", "자산 운용사": "증권·자산운용", "vc": "증권·자산운용",
    "벤처투자사": "증권·자산운용", "투자사": "증권·자산운용", "개인투자법인": "증권·자산운용",
    "액셀러레이터": "증권·자산운용", "사모펀드": "증권·자산운용", "패밀리오피스": "증권·자산운용",
    "벤처캐피탈": "증권·자산운용",
    "연기금": "연기금", "기업연금": "연기금", "공제회": "연기금",
    "보험사": "보험", "은행": "은행",
    "건설사": "건설·엔지니어링", "제조업": "기타 제조",
    "바이오": "제약·바이오", "제약 (글로벌)": "제약·바이오", "소기업 (제약)": "제약·바이오",
    "기업 (생명과학)": "제약·바이오", "화장품 제조업": "화장품·뷰티", "it": "IT·소프트웨어",
    # 비업종 표기 — 거래소/시장 보드·투자 라운드·기업 유형·잡값.
    "biz": None, "ir": None, "기업": None, "기타": None, "벤처기업": None, "스타트업": None,
    "rnd": None, "innovation_growth": None, "금융": None,
    "창업판(차이넥스트)": None, "과창판(스타마켓)": None, "홍콩 증권 거래소(hkex)": None,
    "프라임 스탠다드": None, "series a": None, "series b": None, "series c": None,
}


def _read_pairs() -> dict[str, tuple[str, str]]:
    """엑셀 전체를 읽어 canonical_key → (구분, 회사명)."""
    from openpyxl import load_workbook

    pairs: dict[str, tuple[str, str]] = {}
    for p in sorted(q for d in DIRS if d.is_dir() for q in d.glob("*.xlsx")
                    if not q.name.startswith("~$")):
        try:
            wb = load_workbook(p, read_only=True, data_only=True)
        except Exception as exc:
            print(f"[gubun] 파일 스킵 {p.name}: {exc}", flush=True)
            continue
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            try:
                headers = [_ALIASES.get(str(h).strip() if h else "",
                                        str(h).strip() if h else "") for h in next(rows)]
            except StopIteration:
                continue
            if "gubun" not in headers:
                continue
            blanks = 0
            for raw in rows:
                cells = ["" if v is None else str(v).strip() for v in raw]
                if not any(cells):
                    blanks += 1
                    if blanks > 200:  # importer 와 동일 — 서식만 남은 꼬리 차단.
                        break
                    continue
                blanks = 0
                row = {headers[i]: cells[i] for i in range(min(len(headers), len(cells)))}
                name, dom, gubun = row.get("name", ""), row.get("domain", "") or None, \
                    row.get("gubun", "")
                if not gubun or (not name and not dom):
                    continue
                try:
                    key = canonical_key(domain=dom, name=name, country=row.get("country", ""))
                except ValueError:
                    continue
                pairs.setdefault(key, (gubun, name))
        wb.close()
    return pairs


def main() -> int:
    ap = argparse.ArgumentParser(description="import 행 업종 백필 — 엑셀 구분값 매핑")
    ap.add_argument("--limit-values", type=int, default=0, help="처리할 표기 수 상한(시험용)")
    ap.add_argument("--no-llm", action="store_true", help="수동 별칭·동일표기만 적용(무과금)")
    args = ap.parse_args()

    settings = Settings(industry_llm_max_calls=50_000)  # 기본 캡 5000 < 표기 종수 방어.
    if settings.dry_run:
        print("DRY_RUN=true — 과금 매핑 무의미. 중단.", flush=True)
        return 1
    taxonomy = {t.lower(): t for t in INDUSTRY_TAXONOMY}
    classifier = None if args.no_llm else build_classifier(
        settings, ledger=CostLedger(settings, persist=True)
    )

    pairs = _read_pairs()
    print(f"[gubun] 엑셀 구분 보유 고유 key {len(pairs):,}", flush=True)

    sm = get_sessionmaker(settings)
    with sm() as s:
        target = {k for (k,) in s.execute(text(
            "select d.canonical_key from discovered_company d"
            " left join company co on co.canonical_key = d.canonical_key"
            " where d.source='import' and co.id is null"
            " and coalesce(d.industry,'') in ('', '미분류', '기타 제조')"
        ))}
        # 표기값 → (대상 key 목록, 대표 회사명 샘플).
        by_val: dict[str, list[str]] = defaultdict(list)
        samples: dict[str, list[str]] = defaultdict(list)
        for k, (g, name) in pairs.items():
            if k in target:
                by_val[g].append(k)
                if len(samples[g]) < 3 and name:
                    samples[g].append(name)
        order = sorted(by_val, key=lambda v: -len(by_val[v]))  # 큰 표기 먼저(예산 소진 대비).
        if args.limit_values:
            order = order[: args.limit_values]
        print(f"[gubun] 대상 행 {sum(len(by_val[v]) for v in order):,} · 표기 {len(order):,}종",
              flush=True)

        applied = llm_calls = starved = 0
        for i, val in enumerate(order, start=1):
            low = val.strip().lower()
            if low in _HAND:
                label = _HAND[low]
            elif low in taxonomy:
                label = taxonomy[low]
            elif classifier is None:
                continue
            else:
                verdict = classifier.classify(
                    name=f"업종 구분: {val}",
                    domain=None,
                    text="기존 리드 DB 의 업종 구분 표기다. 이 구분에 속한 대표 회사명: "
                    f"{', '.join(samples[val]) or '(없음)'}. 표기가 업종이 아니라 거래소·"
                    "시장 보드명, 투자 라운드, 기업 유형(벤처기업 등)이면 판정 불가로 하라.",
                )
                llm_calls += 1
                label = verdict.label  # 닫힌 집합 게이트 통과분만, abstain=None.
                starved = starved + 1 if (label is None and not verdict.billed) else 0
                if starved >= 30:  # 캡/예산 소진 추정 — 재실행이 이어받는다.
                    print("[gubun] LLM 캡/예산 소진 추정 — 조기 종료", flush=True)
                    break
            if not label:
                continue
            res = s.execute(
                text("update discovered_company set industry = :lb"
                     " where canonical_key = any(:keys)"
                     " and coalesce(industry,'') in ('', '미분류', '기타 제조')"),
                {"lb": label, "keys": by_val[val]},
            )
            applied += res.rowcount or 0
            if i % 50 == 0:
                s.commit()  # 중간 커밋 — 중단 시 지출·적용분 보존.
                print(f"[gubun] {i}/{len(order)} 표기 | 적용 {applied:,} | LLM {llm_calls:,}",
                      flush=True)
        s.commit()
        print(f"[gubun] 완료 — 적용 {applied:,}행 · LLM {llm_calls:,}콜"
              f" (나머지는 비업종 표기/abstain — 행당 스크립트·승격 분류가 후속)", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
