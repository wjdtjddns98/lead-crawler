"""엑셀 서식 build_row / export 규칙 테스트 (PO 확정)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from leadcrawler.excel_format import FORM_ONLY_NOTE, HEADERS, build_row
from leadcrawler.models import (
    Company,
    CompanyLead,
    Contact,
    ContactType,
    EmailRole,
    EmailValidation,
    ValidationStatus,
)
from leadcrawler.storage.export import ExcelExporter


def _company(**kw: object) -> Company:
    base = dict(canonical_key="dom:x.com", name="테스트기업", country="KR",
                industry="건설", domain="x.com", site_alive=True)
    base.update(kw)
    return Company(**base)


def _full_lead() -> CompanyLead:
    return CompanyLead(
        company=_company(),
        email=Contact(type=ContactType.EMAIL, value="ir@x.com", role=EmailRole.IR),
        phone=Contact(type=ContactType.PHONE, value="+82-2-1-2"),
        form=Contact(type=ContactType.FORM, value="https://x.com/contact"),
        email_validation=EmailValidation(status=ValidationStatus.VALID),
    )


def test_row_rules_full() -> None:
    row = build_row(_full_lead())
    assert row[0] == "KR"  # 국가
    assert row[1] == "테스트기업"  # 업체명
    assert row[3] == "ir@x.com"  # 이메일
    assert row[4] == "O"  # 홈페이지 문의(폼 있음)
    assert row[6] == "" and row[7] == "" and row[11] == ""  # G·H·L 공란
    assert row[8] == "건설"  # 구분=업종만
    assert row[9] == "O"  # 이메일 실존(valid)
    assert row[10] == "O"  # 사이트 실존


def test_form_only_note() -> None:
    lead = CompanyLead(
        company=_company(),
        form=Contact(type=ContactType.FORM, value="https://x.com/contact"),
    )
    row = build_row(lead)
    assert row[3] == ""  # 이메일 없음
    assert row[9] == FORM_ONLY_NOTE  # J = "사이트 내 문의폼"


def test_export_writes_headers(tmp_path: Path) -> None:
    out = ExcelExporter().export([_full_lead()], tmp_path / "out.xlsx")
    wb = load_workbook(out)
    ws = wb.active
    assert [c.value for c in ws[1]] == HEADERS
    assert ws.max_row == 2
