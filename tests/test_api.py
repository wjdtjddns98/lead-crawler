"""검증 웹앱 API 테스트 — 큐 조회/확정/거부/export (in-process, 네트워크 0).

``fastapi`` 미설치(선택적 extra) 면 전체 모듈을 스킵한다.
"""

from __future__ import annotations

import pytest

pytest.importorskip("fastapi")

from fastapi.testclient import TestClient  # noqa: E402

from leadcrawler.config import get_settings  # noqa: E402
from leadcrawler.models import (  # noqa: E402
    Company,
    CompanyLead,
    Contact,
    ContactType,
    EmailRole,
    EmailValidation,
    Listed,
    ValidationStatus,
)
from leadcrawler.security import create_user  # noqa: E402
from leadcrawler.storage.db import init_db, session_scope  # noqa: E402
from leadcrawler.storage.repository import save_lead  # noqa: E402

_USER = "심사원"
_PW = "s3cret-pw"


def _seed(settings) -> None:
    lead = CompanyLead(
        company=Company(
            canonical_key="dom:acme.com",
            name="아크메",
            country="KR",
            industry="건설",
            domain="acme.com",
            homepage="https://acme.com",
            is_active=True,
            site_alive=True,
        ),
        email=Contact(type=ContactType.EMAIL, value="ir@acme.com", role=EmailRole.IR),
        email_validation=EmailValidation(status=ValidationStatus.VALID, mx=True, smtp=True),
    )
    with session_scope(settings) as s:
        save_lead(s, lead, source="test")
        create_user(s, _USER, _PW)


@pytest.fixture
def anon(tmp_path, monkeypatch) -> TestClient:
    """미인증 클라이언트(토큰 없음)."""
    monkeypatch.setenv("LEADCRAWLER_DATABASE_URL", f"sqlite:///{tmp_path}/api.db")
    get_settings.cache_clear()
    settings = get_settings()
    init_db(settings)
    _seed(settings)
    from leadcrawler.api.app import create_app

    return TestClient(create_app())


@pytest.fixture
def client(anon: TestClient) -> TestClient:
    """로그인된 클라이언트 — Authorization 헤더 기본 부착."""
    r = anon.post("/auth/login", json={"username": _USER, "password": _PW})
    assert r.status_code == 200
    anon.headers.update({"Authorization": f"Bearer {r.json()['token']}"})
    return anon


def test_health(client: TestClient) -> None:
    r = client.get("/health")
    assert r.status_code == 200
    assert r.json()["status"] == "ok"


def test_queue_list(client: TestClient) -> None:
    r = client.get("/queue")
    assert r.status_code == 200
    body = r.json()
    assert body["total"] == 1
    item = body["items"][0]
    assert [c["value"] for c in item["candidates"]] == ["ir@acme.com"]
    assert item["selected"] == "ir@acme.com"
    assert item["status"] == "pending"
    assert item["name"] == "아크메"
    assert item["email_status"] == "valid"


def test_queue_status_filter(client: TestClient) -> None:
    assert client.get("/queue", params={"status": "pending"}).json()["total"] == 1
    assert client.get("/queue", params={"status": "confirmed"}).json()["total"] == 0


def test_get_item_and_404(client: TestClient) -> None:
    rid = client.get("/queue").json()["items"][0]["id"]
    assert client.get(f"/queue/{rid}").json()["id"] == rid
    assert client.get("/queue/r_missing").status_code == 404


def test_confirm_then_export(client: TestClient) -> None:
    rid = client.get("/queue").json()["items"][0]["id"]
    r = client.post(f"/queue/{rid}/confirm")
    assert r.status_code == 200
    assert r.json()["status"] == "confirmed"
    assert r.json()["assignee"] == _USER  # 담당자=로그인 사용자(본문 무관).
    # 확정 후 export — 12컬럼 xlsx 가 내려와야 함.
    ex = client.get("/export")
    assert ex.status_code == 200
    assert "spreadsheetml" in ex.headers["content-type"]
    assert len(ex.content) > 0


def test_export_filter_by_country_industry(client: TestClient) -> None:
    import io

    from openpyxl import load_workbook

    rid = client.get("/queue").json()["items"][0]["id"]
    client.post(f"/queue/{rid}/confirm")  # 시드 리드(KR/건설) 확정.

    def data_rows(resp) -> int:
        return load_workbook(io.BytesIO(resp.content)).active.max_row - 1  # 헤더 제외.

    # 매칭(KR/건설) → 1행. 별칭(소문자 'kr')으로도 잡혀야 한다.
    assert data_rows(client.get("/export?country=KR&industry=건설")) == 1
    assert data_rows(client.get("/export?country=대한민국")) == 1  # 한글 별칭 매칭.
    # 불일치 국가/업종 → 0행(헤더만).
    assert data_rows(client.get("/export?country=JP")) == 0
    assert data_rows(client.get("/export?industry=반도체")) == 0


def test_export_filter_by_date(client: TestClient) -> None:
    import io
    from datetime import datetime, timezone

    from openpyxl import load_workbook

    from leadcrawler.schema import ReviewQueueRow
    from leadcrawler.storage.db import session_scope

    rid = client.get("/queue").json()["items"][0]["id"]
    client.post(f"/queue/{rid}/confirm")

    def data_rows(resp) -> int:
        assert resp.status_code == 200
        return load_workbook(io.BytesIO(resp.content)).active.max_row - 1  # 헤더 제외.

    def set_reviewed_at(value) -> None:
        with session_scope(get_settings()) as s:
            s.get(ReviewQueueRow, rid).reviewed_at = value

    # KST 2026-07-21 하루 = UTC [07-20 15:00, 07-21 15:00) — 벽시계 의존 없이
    # reviewed_at 을 고정 UTC 값으로 세팅해 경계 4점 + NULL 을 정밀 검증한다.
    day = "date_from=2026-07-21&date_to=2026-07-21"
    for utc_at, expected in [
        (datetime(2026, 7, 20, 14, 59, 59, 999999, tzinfo=timezone.utc), 0),  # 경계 직전
        (datetime(2026, 7, 20, 15, 0, 0, tzinfo=timezone.utc), 1),  # KST 자정(포함)
        (datetime(2026, 7, 21, 14, 59, 59, 999999, tzinfo=timezone.utc), 1),  # 하루 끝(포함)
        (datetime(2026, 7, 21, 15, 0, 0, tzinfo=timezone.utc), 0),  # 다음날 자정(제외)
        (None, 0),  # 구데이터(reviewed_at NULL)는 날짜 필터 시 제외
    ]:
        set_reviewed_at(utc_at)
        assert data_rows(client.get(f"/export?{day}")) == expected, utc_at
    # 단독 지정 — from 만 / to 만.
    set_reviewed_at(datetime(2026, 7, 21, 0, 0, tzinfo=timezone.utc))
    assert data_rows(client.get("/export?date_from=2026-07-21")) == 1
    assert data_rows(client.get("/export?date_to=2026-07-20")) == 0
    # 형식·범위 오류는 전부 422 — 비ISO/콤팩트 표기/극단 연도(Overflow)/역전 범위.
    assert client.get("/export?date_from=21-07-2026").status_code == 422
    assert client.get("/export?date_from=20260721").status_code == 422
    assert client.get("/export?date_from=0001-01-01").status_code == 422
    assert client.get("/export?date_from=2026-07-22&date_to=2026-07-21").status_code == 422


def test_send_preview_and_dry_run(client: TestClient) -> None:
    # 확정 후 발송 미리보기/발송 — email_send_enabled 기본 false 라 dry-run(실발송 0).
    rid = client.get("/queue").json()["items"][0]["id"]
    client.post(f"/queue/{rid}/confirm")
    prev = client.get("/send/preview").json()
    assert prev["recipients"] == 1 and prev["enabled"] is False
    r = client.post("/send", json={"subject": "안녕하세요", "body": "본문입니다"}).json()
    assert r["dry_run"] is True and r["recipients"] == 1 and r["sent"] == 0


def test_send_empty_subject_422(client: TestClient) -> None:
    # 제목/본문은 필수(min_length=1) → 빈 값이면 422.
    assert client.post("/send", json={"subject": "", "body": "x"}).status_code == 422


def test_reject(client: TestClient) -> None:
    rid = client.get("/queue").json()["items"][0]["id"]
    assert client.post(f"/queue/{rid}/reject").json()["status"] == "rejected"


def test_confirm_with_selection(client: TestClient) -> None:
    rid = client.get("/queue").json()["items"][0]["id"]
    r = client.post(f"/queue/{rid}/confirm", json={"selected": "ir@acme.com"})
    assert r.status_code == 200
    assert r.json()["selected"] == "ir@acme.com" and r.json()["status"] == "confirmed"


def test_confirm_edited_email_registers(client: TestClient) -> None:
    # 후보에 없는 '유효한' 이메일은 사람이 직접 수정/입력한 것으로 등록 후 확정된다.
    rid = client.get("/queue").json()["items"][0]["id"]
    r = client.post(f"/queue/{rid}/confirm", json={"selected": "fixed@acme.com"})
    assert r.status_code == 200
    body = r.json()
    assert body["status"] == "confirmed"
    assert body["selected"] == "fixed@acme.com"  # 수정값이 선택으로 기록.
    assert "fixed@acme.com" in [c["value"] for c in body["candidates"]]  # 후보로 등록.
    # 확정 후 export 에도 수정 이메일이 반영(연락처로 등록되므로).
    assert client.get("/export").status_code == 200


def test_confirm_invalid_format_400(client: TestClient) -> None:
    # 이메일 형식이 아니면 등록 거부(400) — 가비지 후보 차단.
    rid = client.get("/queue").json()["items"][0]["id"]
    r = client.post(f"/queue/{rid}/confirm", json={"selected": "not-an-email"})
    assert r.status_code == 400


def test_confirm_missing_404(client: TestClient) -> None:
    assert client.post("/queue/r_missing/confirm").status_code == 404


def test_confirm_with_homepage_update(client: TestClient) -> None:
    # #185: confirm 본문에 homepage 를 실으면 회사 홈페이지가 갱신되고 응답에 반영된다.
    rid = client.get("/queue").json()["items"][0]["id"]
    r = client.post(
        f"/queue/{rid}/confirm", json={"homepage": "https://corrected.example.com/"}
    )
    assert r.status_code == 200
    body = r.json()
    assert body["status"] == "confirmed"
    assert body["homepage"] == "https://corrected.example.com/"
    # 감사 이력에 변경 전/후 값이 남아야 한다(기존 audit 헬퍼 재사용 — set_review_status).
    from sqlalchemy import select

    from leadcrawler.config import get_settings
    from leadcrawler.schema import ReviewAuditRow
    from leadcrawler.storage.db import session_scope

    with session_scope(get_settings()) as s:
        row = s.scalars(
            select(ReviewAuditRow).where(ReviewAuditRow.review_id == rid)
        ).one()
        assert row.homepage_before == "https://acme.com"
        assert row.homepage_after == "https://corrected.example.com/"


def test_confirm_homepage_null_no_change(client: TestClient) -> None:
    # homepage 생략/None = 변경 없음(하위호환) — 기존 값 그대로 응답.
    rid = client.get("/queue").json()["items"][0]["id"]
    r = client.post(f"/queue/{rid}/confirm", json={"selected": "ir@acme.com"})
    assert r.status_code == 200
    assert r.json()["homepage"] == "https://acme.com"


@pytest.mark.parametrize(
    "bad_homepage",
    ["", "not-a-url", "ftp://acme.com", "javascript:alert(1)", "https://" + "a" * 600 + ".com"],
)
def test_confirm_invalid_homepage_422(client: TestClient, bad_homepage: str) -> None:
    # 신뢰불가 입력(homepage) — 스킴 http/https 형식 위반은 422.
    rid = client.get("/queue").json()["items"][0]["id"]
    r = client.post(f"/queue/{rid}/confirm", json={"homepage": bad_homepage})
    assert r.status_code == 422


def test_invalid_status_422(client: TestClient) -> None:
    # 허용되지 않은 상태 필터는 FastAPI 가 422 로 거부(조용한 빈 결과 방지).
    assert client.get("/queue", params={"status": "bogus"}).status_code == 422


# --- #241: confirm 본문 has_form(문의폼 유무 교정) -----------------------

def _audit_row(rid: str):
    from sqlalchemy import select

    from leadcrawler.schema import ReviewAuditRow
    from leadcrawler.storage.db import session_scope

    with session_scope(get_settings()) as s:
        row = s.scalars(
            select(ReviewAuditRow).where(ReviewAuditRow.review_id == rid)
        ).one()
        s.expunge(row)
        return row


def test_confirm_has_form_true_stores_homepage_link(client: TestClient) -> None:
    # 폼 있음 교정인데 실제 폼 URL 미상 → 홈페이지를 진입 링크로 저장(#241 BE 확정안).
    rid = client.get("/queue").json()["items"][0]["id"]
    r = client.post(f"/queue/{rid}/confirm", json={"has_form": True})
    assert r.status_code == 200
    assert r.json()["form"] == "https://acme.com"
    row = _audit_row(rid)
    assert row.form_before is None
    assert row.form_after == "https://acme.com"


def test_confirm_has_form_true_uses_corrected_homepage(client: TestClient) -> None:
    # 같은 요청에서 homepage 교정도 하면 폼 진입 링크는 교정된 홈페이지를 쓴다.
    rid = client.get("/queue").json()["items"][0]["id"]
    r = client.post(
        f"/queue/{rid}/confirm",
        json={"has_form": True, "homepage": "https://corrected.example.com/"},
    )
    assert r.status_code == 200
    assert r.json()["form"] == "https://corrected.example.com/"


def test_confirm_has_form_false_removes_form(client: TestClient) -> None:
    # 폼 없음 교정 → 저장된 폼 연락처 삭제 + 감사 이력에 전/후 기록.
    from leadcrawler.schema import ContactRow
    from leadcrawler.storage.db import session_scope
    from leadcrawler.storage.repository import contact_id_for

    item = client.get("/queue").json()["items"][0]
    rid, cid = item["id"], item["company_id"]
    with session_scope(get_settings()) as s:
        s.add(
            ContactRow(
                id=contact_id_for(cid, "form", "https://acme.com/contact"),
                company_id=cid,
                type="form",
                value="https://acme.com/contact",
                confidence=0.7,
            )
        )
    assert client.get(f"/queue/{rid}").json()["form"] == "https://acme.com/contact"
    r = client.post(f"/queue/{rid}/confirm", json={"has_form": False})
    assert r.status_code == 200
    assert r.json()["form"] is None
    row = _audit_row(rid)
    assert row.form_before == "https://acme.com/contact"
    assert row.form_after is None


def test_confirm_phone_replaces_and_clears(client: TestClient) -> None:
    # 대표 전화 교정: 값 → 크롤/등록처 전화를 사람 입력(manual·1.0)으로 교체, "" → 지움,
    # None → 변경 없음. 숫자 없는 값은 422.
    from leadcrawler.schema import ContactRow
    from leadcrawler.storage.db import session_scope
    from leadcrawler.storage.repository import contact_id_for

    item = client.get("/queue").json()["items"][0]
    rid, cid = item["id"], item["company_id"]
    with session_scope(get_settings()) as s:
        s.add(ContactRow(id=contact_id_for(cid, "phone", "02-1111-2222"), company_id=cid,
                         type="phone", value="02-1111-2222", extract_method="api",
                         confidence=0.9))
    assert client.get(f"/queue/{rid}").json()["phone"] == "02-1111-2222"
    r = client.post(f"/queue/{rid}/confirm", json={"phone": " 02-3333-4444 "})
    assert r.status_code == 200 and r.json()["phone"] == "02-3333-4444"
    with session_scope(get_settings()) as s:
        rows = s.query(ContactRow).filter_by(company_id=cid, type="phone").all()
        assert [(x.value, x.extract_method, x.confidence) for x in rows] == [
            ("02-3333-4444", "manual", 1.0)
        ]
    # 같은 값 재확정·None 은 무변경.
    assert client.post(f"/queue/{rid}/confirm", json={"phone": "02-3333-4444"}).status_code == 200
    assert client.post(f"/queue/{rid}/confirm", json={}).json()["phone"] == "02-3333-4444"
    assert client.post(f"/queue/{rid}/confirm", json={"phone": "없음"}).status_code == 422
    r = client.post(f"/queue/{rid}/confirm", json={"phone": ""})
    assert r.status_code == 200 and r.json()["phone"] is None


def test_confirm_remove_emails_leaves_form_only(client: TestClient) -> None:
    # 실존하지 않는 이메일 삭제 → 후보 비고 감사이력에 삭제 주소 기록(폼만 남는 리드).
    import json

    from leadcrawler.schema import ContactRow
    from leadcrawler.storage.db import session_scope
    from leadcrawler.storage.repository import contact_id_for

    item = client.get("/queue").json()["items"][0]
    rid, cid = item["id"], item["company_id"]
    r = client.post(f"/queue/{rid}/confirm", json={"remove_emails": ["ir@acme.com"]})
    assert r.status_code == 200
    body = r.json()
    assert body["candidates"] == [] and body["selected"] is None
    with session_scope(get_settings()) as s:
        assert s.get(ContactRow, contact_id_for(cid, "email", "ir@acme.com")) is None
    assert json.loads(_audit_row(rid).emails_removed) == ["ir@acme.com"]


def test_reject_remove_emails(client: TestClient) -> None:
    # 거부 경로도 같은 삭제를 지원한다(본문 없이 호출하면 기존대로 상태만 변경).
    from leadcrawler.schema import ContactRow
    from leadcrawler.storage.db import session_scope
    from leadcrawler.storage.repository import contact_id_for

    item = client.get("/queue").json()["items"][0]
    rid, cid = item["id"], item["company_id"]
    r = client.post(f"/queue/{rid}/reject", json={"remove_emails": ["ir@acme.com"]})
    assert r.status_code == 200
    assert r.json()["status"] == "rejected"
    with session_scope(get_settings()) as s:
        assert s.get(ContactRow, contact_id_for(cid, "email", "ir@acme.com")) is None


def test_reject_without_body_still_works(client: TestClient) -> None:
    rid = client.get("/queue").json()["items"][0]["id"]
    assert client.post(f"/queue/{rid}/reject").json()["status"] == "rejected"


def test_confirm_remove_emails_too_long_422(client: TestClient) -> None:
    rid = client.get("/queue").json()["items"][0]["id"]
    r = client.post(f"/queue/{rid}/confirm", json={"remove_emails": ["a" * 321 + "@x.com"]})
    assert r.status_code == 422


def test_confirm_has_form_true_promotes_low_confidence_form(client: TestClient) -> None:
    # 저신뢰 폴백 폼(0.3)이 있는데 사람이 '폼 있음' 확인 → URL 유지 + 사람확인 승격
    # (form_low_confidence 해제 — '사람 확인 필요' 재노출 방지, 교차리뷰 반영).
    from leadcrawler.schema import ContactRow
    from leadcrawler.storage.db import session_scope
    from leadcrawler.storage.repository import contact_id_for

    item = client.get("/queue").json()["items"][0]
    rid, cid = item["id"], item["company_id"]
    with session_scope(get_settings()) as s:
        s.add(
            ContactRow(
                id=contact_id_for(cid, "form", "https://acme.com/contact"),
                company_id=cid,
                type="form",
                value="https://acme.com/contact",
                confidence=0.3,
            )
        )
    assert client.get(f"/queue/{rid}").json()["form_low_confidence"] is True
    r = client.post(f"/queue/{rid}/confirm", json={"has_form": True})
    assert r.status_code == 200
    body = r.json()
    assert body["form"] == "https://acme.com/contact"  # URL 유지(홈페이지로 덮지 않음)
    assert body["form_confidence"] == 1.0
    assert body["form_low_confidence"] is False


def test_confirm_has_form_true_without_homepage_400(client: TestClient) -> None:
    # 홈페이지조차 없으면 저장할 진입 링크가 없다 → 400(조용한 유실 방지).
    from leadcrawler.models import Company, CompanyLead
    from leadcrawler.storage.db import session_scope
    from leadcrawler.storage.repository import save_lead

    with session_scope(get_settings()) as s:
        save_lead(
            s,
            CompanyLead(
                company=Company(
                    canonical_key="name:nohome|kr",
                    name="노홈페이지",
                    country="KR",
                    industry="건설",
                    is_active=True,
                    site_alive=True,
                )
            ),
            source="test",
        )
    items = client.get("/queue").json()["items"]
    rid = next(it["id"] for it in items if it["name"] == "노홈페이지")
    r = client.post(f"/queue/{rid}/confirm", json={"has_form": True})
    assert r.status_code == 400


def test_confirm_has_form_null_no_change(client: TestClient) -> None:
    # has_form 생략/None = 변경 없음(하위호환).
    rid = client.get("/queue").json()["items"][0]["id"]
    r = client.post(f"/queue/{rid}/confirm", json={"selected": "ir@acme.com"})
    assert r.status_code == 200
    assert r.json()["form"] is None


# --- 확정 본문 note(기타 메모 — 문의폼 미발송 사유 등, 엑셀 L 컬럼) ----------


def _export_ws(client: TestClient):
    import io

    from openpyxl import load_workbook

    return load_workbook(io.BytesIO(client.get("/export").content)).active


def test_confirm_note_stored_and_exported(client: TestClient) -> None:
    rid = client.get("/queue").json()["items"][0]["id"]
    r = client.post(f"/queue/{rid}/confirm", json={"note": "문의폼 접수 오류로 미발송"})
    assert r.status_code == 200
    assert r.json()["note"] == "문의폼 접수 오류로 미발송"
    ws = _export_ws(client)
    assert ws.cell(row=1, column=12).value == "기타"
    assert ws.cell(row=2, column=12).value == "문의폼 접수 오류로 미발송"


def test_confirm_note_none_no_change_and_empty_clears(client: TestClient) -> None:
    rid = client.get("/queue").json()["items"][0]["id"]
    client.post(f"/queue/{rid}/confirm", json={"note": "사유"})
    # None(생략)=변경 없음 — 같은 처리자의 재확정은 허용된다.
    assert client.post(f"/queue/{rid}/confirm", json={}).json()["note"] == "사유"
    # 공백뿐인 문자열=지움(트림 후 빈값 → NULL).
    assert client.post(f"/queue/{rid}/confirm", json={"note": " "}).json()["note"] is None


def test_confirm_note_formula_defused_in_export(client: TestClient) -> None:
    # 검수자 자유텍스트도 수식 인젝션 방어 대상(L 컬럼 defuse).
    rid = client.get("/queue").json()["items"][0]["id"]
    client.post(f"/queue/{rid}/confirm", json={"note": '=HYPERLINK("x")'})
    assert _export_ws(client).cell(row=2, column=12).value == '\'=HYPERLINK("x")'


# --- #238: GET /queue 서버 정렬(sort_by/sort_dir) ------------------------

def _seed_more(names: list[str]) -> None:
    from leadcrawler.models import Company, CompanyLead
    from leadcrawler.storage.db import session_scope
    from leadcrawler.storage.repository import save_lead

    with session_scope(get_settings()) as s:
        for n in names:
            save_lead(
                s,
                CompanyLead(
                    company=Company(
                        canonical_key=f"name:{n}|kr",
                        name=n,
                        country="KR",
                        industry="건설",
                        homepage=f"https://{n}.example.com",
                        is_active=True,
                        site_alive=True,
                    )
                ),
                source="test",
            )


def test_queue_sort_by_name(client: TestClient) -> None:
    _seed_more(["가나", "하하"])
    asc = [it["name"] for it in client.get("/queue", params={"sort_by": "name"}).json()["items"]]
    assert asc == sorted(asc)
    desc = [
        it["name"]
        for it in client.get(
            "/queue", params={"sort_by": "name", "sort_dir": "desc"}
        ).json()["items"]
    ]
    assert desc == sorted(desc, reverse=True)


def test_queue_sort_stable_across_pages(client: TestClient) -> None:
    # 페이지를 나눠 받아도 전체 정렬 순서가 이어진다(#238 의 존재 이유).
    _seed_more(["가나", "하하"])
    whole = [
        it["id"]
        for it in client.get("/queue", params={"sort_by": "name", "limit": 200}).json()["items"]
    ]
    paged = []
    for off in range(0, len(whole), 2):
        paged += [
            it["id"]
            for it in client.get(
                "/queue", params={"sort_by": "name", "limit": 2, "offset": off}
            ).json()["items"]
        ]
    assert paged == whole


def test_queue_sort_invalid_key_422(client: TestClient) -> None:
    assert client.get("/queue", params={"sort_by": "bogus"}).status_code == 422
    assert client.get(
        "/queue", params={"sort_by": "name", "sort_dir": "sideways"}
    ).status_code == 422


def test_login_trims_whitespace(anon: TestClient) -> None:
    # QA①: 아이디/비번 앞뒤 공백(복사·모바일 자동완성)이 있어도 로그인 허용.
    r = anon.post("/auth/login", json={"username": f"  {_USER} ", "password": f" {_PW}  "})
    assert r.status_code == 200
    assert r.json()["username"] == _USER


# --- #191: GET /queue/mine status 필터(처리 이력) -----------------------

def test_mine_no_status_is_unchanged(client: TestClient) -> None:
    """status 생략 = 기존 동작(내 점유 pending) — 하위호환."""
    claimed = client.post("/queue/claim").json()
    mine = client.get("/queue/mine").json()
    assert {it["id"] for it in mine} == {it["id"] for it in claimed}
    assert all(it["status"] == "pending" for it in mine)


def test_mine_status_confirmed_returns_my_history(client: TestClient) -> None:
    rid = client.get("/queue").json()["items"][0]["id"]
    client.post(f"/queue/{rid}/confirm")
    # 확정 후엔 기본(pending) 조회에서 빠지고, status=confirmed 이력에 나타난다.
    assert client.get("/queue/mine").json() == []
    hist = client.get("/queue/mine", params={"status": "confirmed"}).json()
    assert [it["id"] for it in hist] == [rid]
    assert hist[0]["status"] == "confirmed"


def test_mine_status_rejected_returns_my_history(client: TestClient) -> None:
    rid = client.get("/queue").json()["items"][0]["id"]
    client.post(f"/queue/{rid}/reject")
    hist = client.get("/queue/mine", params={"status": "rejected"}).json()
    assert [it["id"] for it in hist] == [rid]
    assert client.get("/queue/mine", params={"status": "confirmed"}).json() == []


def test_mine_status_confirmed_ordered_by_reviewed_at_desc(worker_client: TestClient) -> None:
    """여러 건 확정 시 최신순(reviewed_at desc)."""
    items = worker_client.post("/queue/claim").json()
    ordered_ids = [it["id"] for it in items]
    for rid in ordered_ids:  # 순서대로 확정 → reviewed_at 오름차순 생성.
        assert worker_client.post(f"/queue/{rid}/confirm").status_code == 200
    hist = worker_client.get("/queue/mine", params={"status": "confirmed"}).json()
    assert [it["id"] for it in hist] == list(reversed(ordered_ids))


def test_mine_status_other_users_history_not_included(worker_client: TestClient) -> None:
    """타인이 확정한 항목은 내 이력에 안 뜬다(assignee_id 매칭)."""
    items = worker_client.post("/queue/claim").json()
    for it in items:
        worker_client.post(f"/queue/{it['id']}/confirm")
    other = worker_client.post(
        "/auth/login", json={"username": _ADMIN, "password": _PW}
    ).json()
    admin_client = worker_client
    prior_auth = admin_client.headers["Authorization"]
    admin_client.headers.update({"Authorization": f"Bearer {other['token']}"})
    assert admin_client.get("/queue/mine", params={"status": "confirmed"}).json() == []
    admin_client.headers.update({"Authorization": prior_auth})


def test_mine_invalid_status_422(client: TestClient) -> None:
    assert client.get("/queue/mine", params={"status": "bogus"}).status_code == 422


# --- 작업범위 필터(Filtered Claim) — US-5 라우트/스키마 -------------------

_ADMIN = "관리자"
_WORKER = "직원"
_MIXED = [
    ("kr1.com", "KR", "건설", Listed.UNKNOWN),
    ("us1.com", "US", "Finance", Listed.LISTED),
    ("us2.com", "US", "Finance", Listed.UNLISTED),
]


def _seed_mixed(settings) -> None:
    with session_scope(settings) as s:
        for dom, country, industry, listed in _MIXED:
            lead = CompanyLead(
                company=Company(
                    canonical_key=f"dom:{dom}", name=dom, country=country, industry=industry,
                    domain=dom, homepage=f"https://{dom}", is_active=True, site_alive=True,
                    listed=listed,
                ),
                email=Contact(type=ContactType.EMAIL, value=f"ir@{dom}", role=EmailRole.IR),
                email_validation=EmailValidation(status=ValidationStatus.VALID, mx=True),
            )
            save_lead(s, lead, source="test")
        create_user(s, _ADMIN, _PW)  # 첫 계정 = 자동 admin.
        create_user(s, _WORKER, _PW)  # 두번째 = worker(비관리자).


@pytest.fixture
def worker_client(tmp_path, monkeypatch) -> TestClient:
    """혼합 데이터 + worker(비관리자) 로그인 클라이언트."""
    monkeypatch.setenv("LEADCRAWLER_DATABASE_URL", f"sqlite:///{tmp_path}/mixedapi.db")
    get_settings.cache_clear()
    settings = get_settings()
    init_db(settings)
    _seed_mixed(settings)
    from leadcrawler.api.app import create_app

    c = TestClient(create_app())
    r = c.post("/auth/login", json={"username": _WORKER, "password": _PW})
    assert r.status_code == 200
    c.headers.update({"Authorization": f"Bearer {r.json()['token']}"})
    return c


def test_queue_filters_accessible_to_worker(worker_client: TestClient) -> None:
    """직원(worker)도 /queue/filters 200 — admin 라우트는 그대로 403(오염 없음)."""
    r = worker_client.get("/queue/filters")
    assert r.status_code == 200
    body = r.json()
    assert body["listed"] == ["listed", "unlisted", "unknown"]
    assert len(body["countries"]) > 0 and len(body["industries"]) > 0
    # 구분 옵션 = 저장 어휘(구분 택소노미 + 미분류) — 크롤 타깃용 18개 목록이 아니다(#106).
    from leadcrawler.sources.taxonomy import INDUSTRY_TAXONOMY, UNCLASSIFIED

    assert [o["value"] for o in body["industries"]] == [*INDUSTRY_TAXONOMY, UNCLASSIFIED]
    assert all(o["value"] == o["label"] for o in body["industries"])
    # 동일 직원은 admin 옵션 라우트엔 여전히 접근 불가(분리 확인).
    assert worker_client.get("/admin/countries").status_code == 403
    # 재고 집계도 동일 권한(require_user) — 직원 200(뱃지·비활성 렌더의 소비 주체).
    assert worker_client.get("/queue/stock").status_code == 200


def test_queue_stock_counts_folds_country_and_excludes_claimed(client: TestClient) -> None:
    """/queue/stock — (국가×업종×상장) pending·미점유 집계(P0 재고 가시화 계약).

    표기 혼재('대한민국')는 ISO2 로 접히고, 빈 업종은 '미분류' 뱃지로 나오되 그 값
    그대로 /queue 필터에 넣으면 같은 수가 나온다(왕복 계약 — 리뷰 HIGH 회귀 가드).
    점유(claim)된 행은 재고에서 빠진다.
    """
    from leadcrawler.sources.taxonomy import UNCLASSIFIED

    for key, name, country, industry in [
        ("dom:hanguk.co.kr", "한국사", "대한민국", "건설"),  # 표기 혼재 — KR 로 접힘.
        ("dom:mystery.co.kr", "미분류사", "KR", ""),  # 빈 업종 — '미분류' 뱃지.
    ]:
        lead = CompanyLead(
            company=Company(
                canonical_key=key, name=name, country=country, industry=industry,
                domain=key.removeprefix("dom:"),
                homepage=f"https://{key.removeprefix('dom:')}",
                is_active=True, site_alive=True,
            ),
        )
        with session_scope(get_settings()) as s:
            save_lead(s, lead, source="test")

    r = client.get("/queue/stock")
    assert r.status_code == 200
    body = r.json()
    rows = {(x["country"], x["industry"], x["listed"]): x["n"] for x in body["rows"]}
    assert rows == {
        ("KR", "건설", "unknown"): 2,  # 시드 아크메(KR) + 한국사(대한민국) — 별표기 접힘.
        ("KR", UNCLASSIFIED, "unknown"): 1,  # 빈 업종 → '미분류'.
    }
    assert body["total"] == sum(rows.values()) == 3

    # 왕복 계약: 뱃지 값 그대로 /queue 필터에 넣으면 같은 수 — '미분류'=빈 업종 행 매칭.
    q = client.get("/queue", params={"industry": UNCLASSIFIED}).json()
    assert q["total"] == 1

    # 점유하면 재고에서 빠진다(pending·미점유만 집계 — 기본 배치가 3건을 전부 점유).
    assert client.post("/queue/claim", json={}).status_code == 200
    assert client.get("/queue/stock").json()["total"] == 0


def test_export_worker_own_scope_admin_full(worker_client: TestClient) -> None:
    """export 개방(PO 2026-07-14): worker=200 이지만 자기 확정분(assignee_id)만,
    admin=전체(현행 유지). /send 계열은 여전히 admin 전용(403)."""
    import io

    from openpyxl import load_workbook

    def data_rows(resp) -> int:
        assert resp.status_code == 200
        return load_workbook(io.BytesIO(resp.content)).active.max_row - 1  # 헤더 제외.

    # admin 이 미점유 1건 먼저 확정(이후 worker claim 과 충돌 없게 선행).
    admin = TestClient(worker_client.app)
    r = admin.post("/auth/login", json={"username": _ADMIN, "password": _PW})
    admin.headers.update({"Authorization": f"Bearer {r.json()['token']}"})
    items = admin.get("/queue").json()["items"]
    assert len(items) >= 2
    assert admin.post(f"/queue/{items[0]['id']}/confirm").status_code == 200

    # worker 가 잔여를 점유해 1건 확정 → 자기 export 엔 그 1건만.
    claimed = worker_client.post("/queue/claim").json()
    assert worker_client.post(f"/queue/{claimed[0]['id']}/confirm").status_code == 200
    assert data_rows(worker_client.get("/export")) == 1  # 타인(admin) 확정분 미포함.
    assert data_rows(admin.get("/export")) == 2  # admin=전체.

    # 발송은 권한 완화 범위 밖 — worker 여전히 403.
    assert worker_client.get("/send/preview").status_code == 403


def test_claim_with_country_filter_body(worker_client: TestClient) -> None:
    """POST /queue/claim 본문 필터 — US 만 당겨온다."""
    r = worker_client.post("/queue/claim", json={"country": "US"})
    assert r.status_code == 200
    items = r.json()
    assert {it["country"] for it in items} == {"US"} and len(items) == 2


def test_claim_with_listed_filter_body(worker_client: TestClient) -> None:
    """상장 필터(조인) — listed 만 1건(us1)."""
    r = worker_client.post("/queue/claim", json={"listed": "listed"})
    assert r.status_code == 200
    items = r.json()
    assert {it["name"] for it in items} == {"us1.com"}


def test_claim_invalid_listed_422(worker_client: TestClient) -> None:
    """listed 화이트리스트 밖 값은 422(조용한 빈 결과 방지)."""
    assert worker_client.post("/queue/claim", json={"listed": "bogus"}).status_code == 422


def test_claim_empty_body_is_all(worker_client: TestClient) -> None:
    """본문 생략 = 전체(하위호환) — 3건 전부."""
    r = worker_client.post("/queue/claim")
    assert r.status_code == 200 and len(r.json()) == 3


def test_queue_total_reflects_filter(worker_client: TestClient) -> None:
    """GET /queue total 도 필터 반영(잔여건수)."""
    assert worker_client.get("/queue", params={"country": "US"}).json()["total"] == 2
    assert worker_client.get("/queue", params={"listed": "listed"}).json()["total"] == 1
    assert worker_client.get("/queue", params={"country": "미국"}).json()["total"] == 2  # 별칭.
    assert worker_client.get("/queue").json()["total"] == 3  # 빈 필터=전체.


def test_queue_items_include_listed(worker_client: TestClient) -> None:
    """#122: 큐 응답 아이템에 listed(상장여부) 탑재 — 목록·claim·mine 세 라우트 모두."""
    items = worker_client.get("/queue").json()["items"]
    by_name = {it["name"]: it["listed"] for it in items}
    assert by_name == {"kr1.com": "unknown", "us1.com": "listed", "us2.com": "unlisted"}
    claimed = worker_client.post("/queue/claim").json()
    assert {it["name"]: it["listed"] for it in claimed} == by_name
    mine = worker_client.get("/queue/mine").json()
    assert {it["name"]: it["listed"] for it in mine} == by_name


def test_queue_invalid_listed_422(worker_client: TestClient) -> None:
    assert worker_client.get("/queue", params={"listed": "bogus"}).status_code == 422


# --- 인증 ------------------------------------------------------------

def test_health_is_public(anon: TestClient) -> None:
    assert anon.get("/health").status_code == 200  # 헬스체크는 비보호.


def test_protected_routes_401_without_token(anon: TestClient) -> None:
    assert anon.get("/queue").status_code == 401
    assert anon.get("/export").status_code == 401
    assert anon.post("/queue/x/confirm").status_code == 401
    assert anon.get("/auth/me").status_code == 401


def test_login_wrong_password_401(anon: TestClient) -> None:
    r = anon.post("/auth/login", json={"username": _USER, "password": "wrong"})
    assert r.status_code == 401


def test_login_unknown_user_401(anon: TestClient) -> None:
    r = anon.post("/auth/login", json={"username": "ghost", "password": _PW})
    assert r.status_code == 401


def test_login_throttle_429_after_repeated_failures(tmp_path, monkeypatch) -> None:
    """무차별대입: 임계(3) 실패 후 429 + 잠금 중엔 올바른 비밀번호도 429(인증 전 차단)."""
    monkeypatch.setenv("LEADCRAWLER_DATABASE_URL", f"sqlite:///{tmp_path}/throttle.db")
    monkeypatch.setenv("LEADCRAWLER_LOGIN_MAX_FAILURES", "3")
    get_settings.cache_clear()
    settings = get_settings()
    init_db(settings)
    _seed(settings)
    from leadcrawler.api.app import create_app

    c = TestClient(create_app())
    for _ in range(3):  # 3회 실패 → 401(아직 미잠금에서 시도).
        assert c.post("/auth/login", json={"username": _USER, "password": "wrong"}).status_code == 401
    locked = c.post("/auth/login", json={"username": _USER, "password": "wrong"})
    assert locked.status_code == 429
    assert "Retry-After" in locked.headers
    # 잠긴 동안엔 올바른 비밀번호도 429(scrypt 전에 차단).
    assert c.post("/auth/login", json={"username": _USER, "password": _PW}).status_code == 429


def test_login_request_length_cap_422(anon: TestClient) -> None:
    """과대 페이로드 차단 — username>64·password>256 은 422(scrypt 비용 폭증 방지)."""
    assert anon.post(
        "/auth/login", json={"username": "x" * 65, "password": "y"}
    ).status_code == 422
    assert anon.post(
        "/auth/login", json={"username": "ok", "password": "z" * 257}
    ).status_code == 422


def test_bad_token_401(anon: TestClient) -> None:
    anon.headers.update({"Authorization": "Bearer not-a-real-token"})
    assert anon.get("/queue").status_code == 401


def test_me_returns_username(client: TestClient) -> None:
    assert client.get("/auth/me").json()["username"] == _USER


def test_logout_invalidates_token(client: TestClient) -> None:
    assert client.get("/queue").status_code == 200  # 로그인 상태.
    assert client.post("/auth/logout").status_code == 200
    assert client.get("/queue").status_code == 401  # 폐기된 토큰 → 거부.


def test_logout_without_token_ok(anon: TestClient) -> None:
    # 토큰 없이 로그아웃해도 200(멱등·무해).
    assert anon.post("/auth/logout").status_code == 200


def test_expired_session_rejected(anon: TestClient, monkeypatch) -> None:
    # TTL 0 → 즉시 만료(create_session 은 최소 1시간이지만 now 를 미래로 보정).
    from datetime import datetime, timedelta, timezone

    from leadcrawler import security

    r = anon.post("/auth/login", json={"username": _USER, "password": _PW})
    token = r.json()["token"]
    # 검증 시점을 14시간 뒤로 → 기본 TTL(12h) 초과로 만료 처리.
    future = datetime.now(timezone.utc) + timedelta(hours=14)
    monkeypatch.setattr(security, "_utcnow", lambda: future)
    anon.headers.update({"Authorization": f"Bearer {token}"})
    assert anon.get("/queue").status_code == 401


# --- 확정 본문 has_attachment(첨부파일 유무)·manager(담당자, 엑셀 H 컬럼) ----------


def test_confirm_attachment_and_manager_stored_and_exported(client: TestClient) -> None:
    rid = client.get("/queue").json()["items"][0]["id"]
    r = client.post(
        f"/queue/{rid}/confirm", json={"has_attachment": True, "manager": "김담당"}
    )
    assert r.status_code == 200
    assert r.json()["has_attachment"] is True
    assert r.json()["manager"] == "김담당"
    ws = _export_ws(client)
    assert ws.cell(row=1, column=8).value == "담당자"
    assert ws.cell(row=2, column=8).value == "김담당"


def test_confirm_attachment_manager_none_no_change_and_clear(client: TestClient) -> None:
    rid = client.get("/queue").json()["items"][0]["id"]
    client.post(f"/queue/{rid}/confirm", json={"has_attachment": False, "manager": "담당"})
    # None(생략)=변경 없음 — 같은 처리자의 재확정은 허용된다.
    body = client.post(f"/queue/{rid}/confirm", json={}).json()
    assert body["has_attachment"] is False
    assert body["manager"] == "담당"
    # 공백뿐인 담당자=지움(트림 후 빈값 → NULL). 첨부 체크값은 유지된다.
    body = client.post(f"/queue/{rid}/confirm", json={"manager": " "}).json()
    assert body["manager"] is None
    assert body["has_attachment"] is False


def test_confirm_manager_formula_defused_in_export(client: TestClient) -> None:
    # 담당자명도 자유텍스트 — H 컬럼 defuse 대상.
    rid = client.get("/queue").json()["items"][0]["id"]
    client.post(f"/queue/{rid}/confirm", json={"manager": "=CMD()"})
    assert _export_ws(client).cell(row=2, column=8).value == "'=CMD()"


def test_confirm_note_manager_nul_rejected(client: TestClient) -> None:
    # NUL 바이트는 PG 저장 오류(500) 예방 차원에서 422 로 조기 거절.
    rid = client.get("/queue").json()["items"][0]["id"]
    assert client.post(f"/queue/{rid}/confirm", json={"manager": "김\x00담당"}).status_code == 422
    assert client.post(f"/queue/{rid}/confirm", json={"note": "메모\x00"}).status_code == 422
