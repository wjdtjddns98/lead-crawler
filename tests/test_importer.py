"""기존 import (엑셀/CSV → dedup 시드) 테스트."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from leadcrawler.config import get_settings
from leadcrawler.excel_format import HEADERS
from leadcrawler.importer import ExistingImporter, ImportedCompany
from leadcrawler.storage.db import init_db, session_scope
from leadcrawler.storage.repository import (
    load_seen_domains,
    load_seen_keys,
    seed_discovered_from_imports,
)


def _std_row(country: str, name: str, email: str, site: str) -> list[str]:
    """팀 표준 12컬럼 한 행(국가·업체명·…·사이트(F))을 만든다."""
    return [country, name, "", email, "O", site, "", "", "건설", "O", "O", ""]


def test_import_xlsx_round_trip(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    # 국가, 업체명, ..., 사이트(F)
    row = ["KR", "이미찾은기업", "", "ir@found.com", "O", "found.com",
           "", "", "건설", "O", "O", ""]
    ws.append(row)
    path = tmp_path / "existing.xlsx"
    wb.save(path)

    rows = ExistingImporter().read(path)
    assert len(rows) == 1
    assert rows[0].canonical_key == "dom:found.com"
    assert rows[0].name == "이미찾은기업"


def test_import_csv(tmp_path: Path) -> None:
    path = tmp_path / "existing.csv"
    path.write_text("업체명,국가,사이트\nACME,US,acme.com\n", encoding="utf-8")
    rows = ExistingImporter().read(path)
    assert len(rows) == 1
    assert rows[0].canonical_key == "dom:acme.com"


def test_blank_rows_skipped(tmp_path: Path) -> None:
    path = tmp_path / "e.csv"
    path.write_text("업체명,사이트\n,\nACME,acme.com\n", encoding="utf-8")
    assert len(ExistingImporter().read(path)) == 1


def test_unidentifiable_row_is_skipped(tmp_path: Path) -> None:
    # 회사명이 기호/숫자(정규화 시 빈 문자열)이고 도메인도 정규화 불가('없음')인 행은
    # canonical_key 를 못 만들어도 크래시 없이 건너뛴다(실 데이터 robustness).
    path = tmp_path / "junk.csv"
    path.write_text("업체명,사이트\n①,없음\nACME,acme.com\n", encoding="utf-8")
    rows = ExistingImporter().read(path)
    assert [r.canonical_key for r in rows] == ["dom:acme.com"]


def test_import_reads_all_sheets(tmp_path: Path) -> None:
    # 데이터가 국가별 다중 시트에 흩어진 실제 서식 — 활성 시트만 읽으면 대부분 누락된다.
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "대한민국"
    ws1.append(HEADERS)
    ws1.append(_std_row("KR", "가기업", "ir@a.com", "a.com"))
    ws2 = wb.create_sheet("미국")
    ws2.append(HEADERS)
    ws2.append(_std_row("US", "B Corp", "ir@b.com", "b.com"))
    path = tmp_path / "multi.xlsx"
    wb.save(path)

    keys = {r.canonical_key for r in ExistingImporter().read(path)}
    assert keys == {"dom:a.com", "dom:b.com"}  # 활성 시트 1개가 아니라 두 시트 모두.


def test_blank_row_guard_stops_phantom_rows(tmp_path: Path) -> None:
    # read_only 모드가 서식만 있는 빈 행을 대량으로 흘려도 데이터 1건만 반환해야 한다.
    wb = Workbook()
    ws = wb.active
    ws.append(HEADERS)
    ws.append(_std_row("KR", "유일기업", "ir@only.com", "only.com"))
    # 연속 빈 행을 가드 한계보다 많이 추가 — 이후엔 더 읽지 않아야 한다.
    for _ in range(500):
        ws.append([None] * len(HEADERS))
    path = tmp_path / "phantom.xlsx"
    wb.save(path)

    rows = ExistingImporter().read(path)
    assert [r.name for r in rows] == ["유일기업"]


def test_seed_discovered_persists_and_idempotent(tmp_path: Path) -> None:
    # conftest 가 DATABASE_URL 을 격리 SQLite 로 잡아준다 — 스키마만 생성.
    settings = get_settings()
    init_db(settings)
    companies = [
        ImportedCompany(
            canonical_key="dom:seed1.com", name="시드1",
            country="대한민국", domain="https://seed1.com/ir",
        ),
        ImportedCompany(
            canonical_key="name:kr:이름만", name="이름만", country="대한민국", domain=None
        ),
    ]
    with session_scope(settings) as s:
        new, skipped = seed_discovered_from_imports(s, companies)
    assert (new, skipped) == (2, 0)

    with session_scope(settings) as s:
        assert load_seen_keys(s) == {"dom:seed1.com", "name:kr:이름만"}
        # URL 이 eTLD+1 로 정규화 저장돼 도메인 동치 dedup 시드가 된다.
        assert load_seen_domains(s) == {"seed1.com"}

    # 재실행 멱등 — 모두 기존 스킵(식별정보 보존, 재적재 없음).
    with session_scope(settings) as s:
        new2, skipped2 = seed_discovered_from_imports(s, companies)
    assert (new2, skipped2) == (0, 2)
