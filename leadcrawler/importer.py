"""기존 검색분 import — 엑셀/CSV → 중복제거 시드.

제약 ①(이미 검색한 기업 재추출 금지)의 선행 단계. 직원들이 만든 동일 12컬럼
서식(또는 회사명·국가·사이트·이메일 헤더를 가진 CSV)을 읽어 canonical_key 를
산정한 ``ImportedCompany`` 목록을 돌려준다.
"""

from __future__ import annotations

import csv
from collections.abc import Iterator
from pathlib import Path

from pydantic import BaseModel

from .dedup import canonical_key

# 다양한 헤더 표기를 표준 필드로 흡수하기 위한 매핑. 주의: ``홈페이지 문의``는 도메인이
# 아니라 문의폼 유무 표시 컬럼이므로 매핑하지 않는다(별도 ``사이트``가 실제 도메인).
_HEADER_ALIASES = {
    "국가": "country",
    "업체명": "name",
    "회사명": "name",
    "사이트": "domain",
    "홈페이지": "domain",
    "site": "domain",
    "도메인": "domain",
    "이메일": "email",
    "이메일 (IR 우선순위)": "email",
    "email": "email",
}

# read_only 워크북은 서식만 있는 빈 행을 시트 최대행(약 104만)까지 흘릴 수 있다. 연속
# 빈 행이 이 한계를 넘으면 시트 끝으로 간주해 불필요한 100만 행 순회를 차단한다.
_MAX_CONSECUTIVE_BLANKS = 200


class ImportedCompany(BaseModel):
    """기존 데이터 한 행 — 시드용 최소 정보 + canonical_key."""

    canonical_key: str
    name: str
    country: str = ""
    domain: str | None = None
    email: str | None = None


def _normalize_headers(headers: list[str]) -> list[str]:
    return [_HEADER_ALIASES.get((h or "").strip(), (h or "").strip()) for h in headers]


def _row_to_company(row: dict[str, str]) -> ImportedCompany | None:
    name = (row.get("name") or "").strip()
    domain = (row.get("domain") or "").strip() or None
    country = (row.get("country") or "").strip()
    email = (row.get("email") or "").strip() or None
    if not name and not domain:
        return None
    try:
        key = canonical_key(domain=domain, name=name, country=country)
    except ValueError:
        # 도메인이 정규화 불가(예: 'N/A')이고 회사명도 정규화 시 빈 문자열(기호·숫자만)인
        # 행 — 식별 key 를 만들 수 없으므로 안전하게 건너뛴다.
        return None
    return ImportedCompany(
        canonical_key=key, name=name, country=country, domain=domain, email=email
    )


def _iter_csv(path: Path) -> Iterator[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return
    headers = _normalize_headers(rows[0])
    for raw in rows[1:]:
        yield {headers[i]: raw[i] for i in range(min(len(headers), len(raw)))}


def _iter_xlsx(path: Path) -> Iterator[dict[str, str]]:
    """엑셀의 **모든 시트**를 순회한다(국가별 시트에 데이터가 흩어져 있어 활성 시트만
    읽으면 대부분 누락된다). 각 시트는 첫 행을 헤더로 보고, 시트마다 헤더가 다를 수 있어
    시트별로 다시 정규화한다. 연속 빈 행이 과다하면 시트 끝으로 간주하고 중단한다.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows)
            except StopIteration:
                continue
            headers = _normalize_headers(
                [str(h) if h is not None else "" for h in header_row]
            )
            blanks = 0
            for raw in rows:
                cells = ["" if v is None else str(v) for v in raw]
                if not any(c.strip() for c in cells):
                    blanks += 1
                    if blanks >= _MAX_CONSECUTIVE_BLANKS:
                        break  # 빈 행 과다 → 데이터 종료(서식만 남은 100만 행 순회 차단).
                    continue
                blanks = 0
                yield {headers[i]: cells[i] for i in range(min(len(headers), len(cells)))}
    finally:
        wb.close()


class ExistingImporter:
    """엑셀/CSV 기존 데이터를 시드용 ``ImportedCompany`` 목록으로 읽는다."""

    def read(self, path: str | Path) -> list[ImportedCompany]:
        """파일 확장자에 따라 .xlsx/.csv 를 파싱한다(빈/식별불가 행은 건너뜀)."""
        p = Path(path)
        if p.suffix.lower() in {".xlsx", ".xlsm"}:
            rows = _iter_xlsx(p)
        else:
            rows = _iter_csv(p)
        out: list[ImportedCompany] = []
        for row in rows:
            company = _row_to_company(row)
            if company is not None:
                out.append(company)
        return out
