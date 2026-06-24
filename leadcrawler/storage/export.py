"""고정 엑셀 서식 산출 (openpyxl).

검증 완료된 :class:`CompanyLead` 목록을 PO 확정 12컬럼 서식으로 저장한다.
"""

from __future__ import annotations

from pathlib import Path

from ..excel_format import HEADERS, build_row
from ..models import CompanyLead


class ExcelExporter:
    """:class:`CompanyLead` 목록을 12컬럼 .xlsx 로 내보낸다."""

    def export(self, leads: list[CompanyLead], path: str | Path) -> Path:
        """``path`` 에 헤더 + 각 리드 행을 기록하고 경로를 반환한다."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "기업 리스트"
        ws.append(HEADERS)
        for lead in leads:
            ws.append(build_row(lead))
            # E(홈페이지 문의=폼 URL)·F(사이트 URL)가 URL이면 클릭 가능한 하이퍼링크로.
            row = ws.max_row
            for col in (5, 6):
                cell = ws.cell(row=row, column=col)
                value = cell.value
                if isinstance(value, str) and value.startswith(("http://", "https://")):
                    cell.hyperlink = value
                    cell.style = "Hyperlink"

        out = Path(path)
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out)
        return out
